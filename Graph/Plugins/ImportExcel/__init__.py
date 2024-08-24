# Plugin for import Excel .xlsx files.
# This requires openpyxl (https://openpyxl.readthedocs.io/en/stable/), et_xmlfile (https://pypi.org/project/et_xmlfile/) and jdcal (https://pypi.org/project/jdcal/)
import Graph
import vcl
import sys
import os.path
import math
import traceback

PluginName = "Import Excel files"
PluginVersion = "0.3"
PluginDescription = ("Creates a menu item File|Import|Excel file, which can be used to import one or more point series from an .xlsx file. " +
                    "It also adds a menu item Import Excel file to the context menu in the Function|Insert point series dialog, which will import a .xlsx file into the table.")

def ImportExcelFile(FileName, Grid):
    import openpyxl
    wb = openpyxl.load_workbook(FileName, data_only=True)
    ws = wb[wb.sheetnames[0]]
    RowIndex = 1
    ColIndex = 0
    for Row in ws.values:
        for Value in Row[:Grid.ColCount]:
            if type(Value) in [int, float]:
                Grid.Cells[ColIndex, RowIndex] = Graph.FormatNumber(Value, 12, Graph.cfReal)
            elif type(Value) is str:
                Grid.Cells[ColIndex, RowIndex] = Value
            else:
                Grid.Cells[ColIndex, RowIndex] = ""    
            ColIndex += 1
        RowIndex += 1
        ColIndex = 0
    Grid.RowCount = RowIndex + 1
    
def ShowOpenDialog():
    Dialog = vcl.TOpenDialog(None)
    Dialog.Options = "ofHideReadOnly,ofFileMustExist"
    Dialog.Filter = "Microsoft Excel 2007-2013 XML [*.xlsx]|*.xlsx"
    if Dialog.Execute():
        return Dialog.FileName

def MenuClick(Sender):
    FileName = ShowOpenDialog()
    if FileName:
        try:
            ImportExcelFile(FileName, Sender.Owner.Grid)
        except Exception as E:
            traceback.print_exc()
            vcl.Application.MessageBox(E.args[0], "Import Excel file error", 0x10)            

def OnShowForm(Form):
    if Form.Name == "Form14":
        MenuItem = vcl.TMenuItem(Form, Caption="Import Excel file...", OnClick=MenuClick, _owned=False)
        Form.PopupMenu1.Items.Insert(Form.Popup_Import.MenuIndex + 1, MenuItem)

def ImportAll(FileName):
    import openpyxl
    wb = openpyxl.load_workbook(FileName, data_only=True)
    ws = wb[wb.sheetnames[0]]
    xCol = ws["A"]
    SeriesAdded = False
    for Index, Col in enumerate(ws.iter_cols(min_col=2, values_only=True), start=1):
        Series = Graph.TPointSeries()
        Series.Size = 4
        Series.LegendText = os.path.splitext(os.path.split(FileName)[1])[0] + "_" + str(Index)
        for P in zip(xCol, Col):
            if type(P[0].value) in [int, float] and type(P[1]) in [int, float]:
                Series.Points.append((P[0].value, P[1]))
        if len(Series.Points) > 0:
            Graph.FunctionList.append(Series)
            SeriesAdded = True
    if not SeriesAdded:
        vcl.Application.MessageBox("No valid data found in the Excel file.", "Import Excel file error", 0x10)
        
def GlobalImport(Sender):
    FileName = ShowOpenDialog()
    if FileName:
        try:
            ImportAll(FileName)
        except Exception as E:
            print(E.args)
            traceback.print_exc()
            vcl.Application.MessageBox(E.args[0], "Import Excel file error", 0x10)
        
Graph.OnShowForm.append(OnShowForm)
sys.path.append(os.path.dirname(__file__))

Action = Graph.CreateAction("Excel file...", GlobalImport, IconFile="ExcelIcon.png")
Graph.AddActionToMainMenu(Action, "File", [Graph.GetText("Import")])
